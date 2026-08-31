#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""수출서류 템플릿(xlsx) 생성.

forms/수출서류_상업송장_포장명세서.xlsx 를 만든다.
[입력] 시트 하나만 채우면 Commercial Invoice 와 Packing List 두 시트가 완성되는 구조.

    python3 automation/build_template.py

레이아웃을 고칠 때는 이 스크립트를 고치고 다시 돌릴 것. xlsx 를 직접 편집하지 말 것.
generate.py 가 이 템플릿의 [입력] 시트 좌표에 값을 써넣으므로 좌표를 바꾸면 그쪽도 고쳐야 한다.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, 'forms', '수출서류_상업송장_포장명세서.xlsx')

품목_시작, 품목_끝 = 43, 57            # [입력] 시트의 품목 행 범위
품목_최대 = 품목_끝 - 품목_시작 + 1     # 15행

얇은 = Side(style='thin', color='000000')
굵은 = Side(style='medium', color='000000')
BOX = Border(left=얇은, right=얇은, top=얇은, bottom=얇은)
노랑 = PatternFill('solid', start_color='FFF9DB', end_color='FFF9DB')
회색 = PatternFill('solid', start_color='EFEFEF', end_color='EFEFEF')

A = lambda **kw: Alignment(**kw)
왼쪽 = A(horizontal='left', vertical='center', wrap_text=False)
줄바꿈 = A(horizontal='left', vertical='center', wrap_text=True)   # 긴 품명이 잘리지 않게
가운데 = A(horizontal='center', vertical='center')
오른쪽 = A(horizontal='right', vertical='center')

수량fmt, 금액fmt, 중량fmt = '#,##0', '#,##0.00', '#,##0.00'


def 셀(ws, addr, value=None, *, bold=False, size=10, align=왼쪽, fill=None,
       border=True, fmt=None, merge=None):
    """값/서식을 한 번에. merge 를 주면 그 범위를 병합하고 테두리를 범위 전체에 두른다."""
    c = ws[addr]
    c.value = value
    c.font = Font(name='Arial', size=size, bold=bold)
    c.alignment = align
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    if merge:
        ws.merge_cells(merge)
        for row in ws[merge]:
            for cc in row:
                if border:
                    cc.border = BOX
                if fill:
                    cc.fill = fill
    elif border:
        c.border = BOX
    return c


def 라벨값(ws, row, 라벨1, 값1, 라벨2=None, 값2=None):
    """A:B 라벨 / C:D 값 / E:F 라벨 / G:H 값 형태의 한 줄."""
    셀(ws, f'A{row}', 라벨1, bold=True, fill=회색, merge=f'A{row}:B{row}')
    셀(ws, f'C{row}', 값1, merge=f'C{row}:D{row}')
    if 라벨2 is not None:
        셀(ws, f'E{row}', 라벨2, bold=True, fill=회색, merge=f'E{row}:F{row}')
        셀(ws, f'G{row}', 값2, merge=f'G{row}:H{row}')


def 인쇄설정(ws, 범위, 가로=False):
    ws.print_area = 범위
    ws.page_setup.orientation = 'landscape' if 가로 else 'portrait'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.5
    ws.page_margins.top = ws.page_margins.bottom = 0.6
    ws.print_options.horizontalCentered = True


# ══════════════════════════════════════════════════════════════
# 1. [입력] 시트
# ══════════════════════════════════════════════════════════════
def 입력시트(wb):
    ws = wb.create_sheet('입력')
    for col, w in zip('ABCDEFGHIJK', [22, 34, 46, 12, 10, 8, 12, 14, 14, 14, 10]):
        ws.column_dimensions[col].width = w

    셀(ws, 'A1', '수출서류 입력시트', bold=True, size=14, border=False)
    셀(ws, 'A2', '노란색 칸만 채우면 Commercial Invoice 와 Packing List 두 시트가 자동으로 완성됨. '
                 '흰색 칸은 수식이므로 건드리지 말 것.', border=False)

    def 구역(row, 제목):
        셀(ws, f'A{row}', 제목, bold=True, fill=회색, border=False)

    def 항목(row, 라벨, 예시, 힌트=''):
        셀(ws, f'A{row}', 라벨, bold=True, border=False)
        셀(ws, f'B{row}', 예시, fill=노랑)
        셀(ws, f'C{row}', 힌트, size=9, border=False).font = Font(name='Arial', size=9, color='808080')

    구역(4, '■ 거래 기본정보')
    for r, (라, 예, 힌) in enumerate([
        ('Invoice No.',            'SY-2026-0001', '송장번호. 회사 규칙대로'),
        ('Invoice Date',           '2026-08-31',   'YYYY-MM-DD'),
        ('P/O No.',                'PO-8842',      '바이어 주문번호. 없으면 공란'),
        ('가격조건 (Incoterms)',    'FOB BUSAN',    '반드시 지정장소까지. 예) FOB BUSAN'),
        ('Incoterms 버전',          'Incoterms(R) 2020', '2020 으로 통일'),
        ('결제조건 (Payment)',      'T/T 30% in advance, 70% after B/L', '구체적으로'),
        ('결제통화 (Currency)',     'USD',          'USD / EUR / CNY 등'),
        ('원산지 (Country of Origin)', 'REPUBLIC OF KOREA', ''),
        ('선적항 (Port of Loading)',   'BUSAN, KOREA', ''),
        ('도착항 (Port of Discharge)', 'HO CHI MINH, VIETNAM', ''),
        ('최종목적지 (Final Destination)', 'HO CHI MINH, VIETNAM', ''),
        ('선박명/항차 (Vessel/Voyage)',   'T.B.A.', '미정이면 T.B.A.'),
        ('출항예정일 (ETD)',        'T.B.A.',       '미정이면 T.B.A.'),
    ], start=5):
        항목(r, 라, 예, 힌)

    구역(19, '■ 수출자 (Shipper / Exporter)')
    for r, (라, 예, 힌) in enumerate([
        ('상호',        'SY COSMETICS CO., LTD.',      '통관고유부호 등록 상호와 동일하게'),
        ('주소1',       '123, TEHERAN-RO, GANGNAM-GU', ''),
        ('주소2',       'SEOUL, 06234, REPUBLIC OF KOREA', ''),
        ('담당자/연락처', 'J.S. LEE / +82-2-1234-5678',  ''),
        ('통관고유부호',  '1234567890123',               '유니패스 발급'),
    ], start=20):
        항목(r, 라, 예, 힌)

    구역(26, '■ 수입자 (Consignee)')
    for r, (라, 예, 힌) in enumerate([
        ('상호',        'ABC TRADING CO., LTD.',           '해외거래처부호 등록 상호와 동일하게'),
        ('주소1',       '456 NGUYEN HUE STREET, DISTRICT 1', ''),
        ('주소2',       'HO CHI MINH CITY, VIETNAM',       ''),
        ('담당자/연락처', 'MR. TRAN / +84-28-1234-5678',     ''),
        ('해외거래처부호', 'VN1234567',                      '유니패스 등록'),
    ], start=27):
        항목(r, 라, 예, 힌)

    구역(33, '■ 착화통지처 (Notify Party)')
    항목(34, '표기', 'SAME AS CONSIGNEE', '수입자와 같으면 SAME AS CONSIGNEE')

    구역(36, '■ 포장 총괄')
    항목(37, '포장단위',    'CTNS', 'CTNS(박스) / PLTS(팔레트) 등')
    항목(38, '총 부피 (CBM)', 2.45,  '숫자만')
    ws['B38'].number_format = 중량fmt

    구역(41, f'■ 품목 (아래 표만 채우면 됨. 최대 {품목_최대}행)')
    헤더 = ['No.', '품명 (Description)', '규격 (Spec)', 'HS CODE', '수량', '단위',
            '단가', '금액', '순중량 N.W.(kg)', '총중량 G.W.(kg)', '포장수량']
    for i, h in enumerate(헤더):
        셀(ws, f'{get_column_letter(i+1)}42', h, bold=True, align=가운데, fill=회색)

    예시 = [1, 'FACIAL CLEANSING FOAM 150ML', 'MODEL SY-CF150', '3304.99',
            1200, 'PCS', 3.5, None, 180, 210.5, 50]
    for r in range(품목_시작, 품목_끝 + 1):
        for i in range(11):
            col = get_column_letter(i + 1)
            v = 예시[i] if r == 품목_시작 else None
            fmt = {4: 수량fmt, 6: 금액fmt, 7: 금액fmt, 8: 중량fmt, 9: 중량fmt, 10: 수량fmt}.get(i)
            al = 오른쪽 if i in (0, 4, 6, 7, 8, 9, 10) else 왼쪽
            if col == 'H':                       # 금액 = 수량 x 단가
                셀(ws, f'H{r}', f'=IF(E{r}="","",E{r}*G{r})', align=오른쪽, fmt=금액fmt)
            else:
                셀(ws, f'{col}{r}', v, align=al, fill=노랑, fmt=fmt)

    셀(ws, f'A{품목_끝+1}', 'TOTAL', bold=True, align=가운데, fill=회색, merge=f'A{품목_끝+1}:D{품목_끝+1}')
    for col, fmt in [('E', 수량fmt), ('H', 금액fmt), ('I', 중량fmt), ('J', 중량fmt), ('K', 수량fmt)]:
        셀(ws, f'{col}{품목_끝+1}', f'=SUM({col}{품목_시작}:{col}{품목_끝})',
          bold=True, align=오른쪽, fill=회색, fmt=fmt)
    for col in 'FG':
        셀(ws, f'{col}{품목_끝+1}', None, fill=회색)

    셀(ws, 'A60', '작성 주의', bold=True, border=False)
    for i, t in enumerate([
        '01  품명은 관세사가 HS 분류를 할 수 있을 만큼 구체적으로. PARTS 같은 표기는 분류 불가.',
        '02  수출자/수입자 상호와 주소는 통관고유부호 및 해외거래처부호 등록정보와 대소문자까지 일치시킬 것.',
        '03  원산지증명서를 발급할 예정이면 이 일치 여부가 특히 중요함. 불일치 시 C/O 재발행 대상.',
        '04  총중량(G.W.)은 포장 포함, 순중량(N.W.)은 포장 제외. 신고중량과 실제중량 차이는 미선적의 주원인.',
        '05  CIF/CFR 조건이면 운임과 보험료 금액을 관세사에게 별도로 알려주어야 수출신고가 됨.',
    ], start=61):
        셀(ws, f'A{i}', t, size=9, border=False)
    인쇄설정(ws, 'A1:K65', 가로=True)     # 한 장에 들어가게. PDF 분리 시 페이지 수를 예측 가능하게 함
    return ws


# ══════════════════════════════════════════════════════════════
# 2. 공통 머리부 (C/I, P/L 이 같은 모양)
# ══════════════════════════════════════════════════════════════
def 머리부(ws, 제목, 마지막열):
    셀(ws, 'A1', 제목, bold=True, size=16, align=가운데, border=False,
      merge=f'A1:{마지막열}1')
    ws.row_dimensions[1].height = 26

    셀(ws, 'A3', 'Shipper / Exporter', bold=True, fill=회색, merge='A3:D3')
    for i, src in enumerate(['B20', 'B21', 'B22', 'B23']):
        셀(ws, f'A{4+i}', f'=입력!{src}', merge=f'A{4+i}:D{4+i}')

    for i, (라, src) in enumerate([('Invoice No.', 'B5'), ('Date', 'B6'), ('P/O No.', 'B7')]):
        셀(ws, f'E{3+i}', 라, bold=True, fill=회색, merge=f'E{3+i}:F{3+i}')
        셀(ws, f'G{3+i}', f'=입력!{src}', merge=f'G{3+i}:H{3+i}')
    for r in (6, 7):                                   # 오른쪽 여백 칸도 테두리 유지
        셀(ws, f'E{r}', None, merge=f'E{r}:H{r}')

    셀(ws, 'A8', 'Consignee',    bold=True, fill=회색, merge='A8:D8')
    셀(ws, 'E8', 'Notify Party', bold=True, fill=회색, merge='E8:H8')
    for i, src in enumerate(['B27', 'B28', 'B29', 'B30']):
        셀(ws, f'A{9+i}', f'=입력!{src}', merge=f'A{9+i}:D{9+i}')
    셀(ws, 'E9', '=입력!B34', merge='E9:H9')
    for r in (10, 11, 12):
        셀(ws, f'E{r}', None, merge=f'E{r}:H{r}')

    for i, (라1, s1, 라2, s2) in enumerate([
        ('Port of Loading',   'B13', 'Port of Discharge', 'B14'),
        ('Final Destination', 'B15', 'Vessel / Voyage',   'B16'),
        ('ETD',               'B17', 'Country of Origin', 'B12'),
        ('Price Term',        'B8',  'Incoterms',         'B9'),
        ('Payment Term',      'B10', 'Currency',          'B11'),
    ]):
        라벨값(ws, 13 + i, 라1, f'=입력!{s1}', 라2, f'=입력!{s2}')


def 서명부(ws, row, 마지막열):
    셀(ws, f'F{row}',   '=입력!B20', bold=True, align=가운데, border=False, merge=f'F{row}:{마지막열}{row}')
    셀(ws, f'F{row+1}', 'Signed by', align=가운데, border=False, merge=f'F{row+1}:{마지막열}{row+1}')
    셀(ws, f'F{row+2}', '________________________', align=가운데, border=False,
      merge=f'F{row+2}:{마지막열}{row+2}')


# ══════════════════════════════════════════════════════════════
# 3. Commercial Invoice
# ══════════════════════════════════════════════════════════════
def 상업송장(wb):
    ws = wb.create_sheet('Commercial Invoice')
    for col, w in zip('ABCDEFGH', [6, 32, 18, 12, 11, 7, 13, 15]):
        ws.column_dimensions[col].width = w
    머리부(ws, 'COMMERCIAL INVOICE', 'H')

    헤더 = ["No.", "Description", "Spec", "HS CODE", "Q'ty", "Unit", "Unit Price", "Amount"]
    for i, h in enumerate(헤더):
        셀(ws, f'{get_column_letter(i+1)}19', h, bold=True, align=가운데, fill=회색)

    for k in range(품목_최대):
        r, s = 20 + k, 품목_시작 + k
        for i, (col, src) in enumerate(zip('ABCDEFGH', 'ABCDEFGH')):
            fmt = {'E': 수량fmt, 'G': 금액fmt, 'H': 금액fmt}.get(col)
            al = 오른쪽 if col in 'AEGH' else (가운데 if col in 'DF' else
                 (줄바꿈 if col in 'BC' else 왼쪽))
            셀(ws, f'{col}{r}', f'=IF(입력!$B${s}="","",입력!{src}{s})', align=al, fmt=fmt)

    합계행 = 20 + 품목_최대
    셀(ws, f'A{합계행}', 'TOTAL', bold=True, align=가운데, fill=회색, merge=f'A{합계행}:D{합계행}')
    셀(ws, f'E{합계행}', f'=입력!E{품목_끝+1}', bold=True, align=오른쪽, fill=회색, fmt=수량fmt)
    for col in 'FG':
        셀(ws, f'{col}{합계행}', None, fill=회색)
    셀(ws, f'H{합계행}', f'=입력!H{품목_끝+1}', bold=True, align=오른쪽, fill=회색, fmt=금액fmt)

    r = 합계행 + 2
    셀(ws, f'A{r}', 'Total Amount', bold=True, fill=회색, merge=f'A{r}:B{r}')
    셀(ws, f'C{r}', f'=입력!B11 & " " & TEXT(입력!H{품목_끝+1},"#,##0.00")',
      bold=True, merge=f'C{r}:H{r}')
    셀(ws, f'A{r+1}', 'Price Term', bold=True, fill=회색, merge=f'A{r+1}:B{r+1}')
    셀(ws, f'C{r+1}', '=입력!B8 & " " & 입력!B9', merge=f'C{r+1}:H{r+1}')

    서명부(ws, r + 3, 'H')
    인쇄설정(ws, f'A1:H{r+5}')
    return ws


# ══════════════════════════════════════════════════════════════
# 4. Packing List
# ══════════════════════════════════════════════════════════════
def 포장명세서(wb):
    ws = wb.create_sheet('Packing List')
    for col, w in zip('ABCDEFGH', [6, 32, 18, 12, 11, 7, 13, 15]):
        ws.column_dimensions[col].width = w
    머리부(ws, 'PACKING LIST', 'H')

    헤더 = ["No.", "Description", "Spec", "HS CODE", "Q'ty", "Unit", "N.W.(kg)", "G.W.(kg)"]
    for i, h in enumerate(헤더):
        셀(ws, f'{get_column_letter(i+1)}19', h, bold=True, align=가운데, fill=회색)

    # C/I 의 단가·금액 자리에 순중량·총중량이 들어간다 ([입력] I열, J열)
    맵 = list(zip('ABCDEFGH', ['A', 'B', 'C', 'D', 'E', 'F', 'I', 'J']))
    for k in range(품목_최대):
        r, s = 20 + k, 품목_시작 + k
        for col, src in 맵:
            fmt = {'E': 수량fmt, 'G': 중량fmt, 'H': 중량fmt}.get(col)
            al = 오른쪽 if col in 'AEGH' else (가운데 if col in 'DF' else
                 (줄바꿈 if col in 'BC' else 왼쪽))
            셀(ws, f'{col}{r}', f'=IF(입력!$B${s}="","",입력!{src}{s})', align=al, fmt=fmt)

    합계행 = 20 + 품목_최대
    셀(ws, f'A{합계행}', 'TOTAL', bold=True, align=가운데, fill=회색, merge=f'A{합계행}:D{합계행}')
    셀(ws, f'E{합계행}', f'=입력!E{품목_끝+1}', bold=True, align=오른쪽, fill=회색, fmt=수량fmt)
    for col in 'F':
        셀(ws, f'{col}{합계행}', None, fill=회색)
    셀(ws, f'G{합계행}', f'=입력!I{품목_끝+1}', bold=True, align=오른쪽, fill=회색, fmt=중량fmt)
    셀(ws, f'H{합계행}', f'=입력!J{품목_끝+1}', bold=True, align=오른쪽, fill=회색, fmt=중량fmt)

    r = 합계행 + 2
    셀(ws, f'A{r}', 'Total Packages', bold=True, fill=회색, merge=f'A{r}:B{r}')
    셀(ws, f'C{r}', f'=입력!K{품목_끝+1} & " " & 입력!B37', bold=True, merge=f'C{r}:H{r}')
    셀(ws, f'A{r+1}', 'Total Measurement', bold=True, fill=회색, merge=f'A{r+1}:B{r+1}')
    셀(ws, f'C{r+1}', f'=TEXT(입력!B38,"#,##0.00") & " CBM"', merge=f'C{r+1}:H{r+1}')

    서명부(ws, r + 3, 'H')
    인쇄설정(ws, f'A1:H{r+5}')
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)
    입력시트(wb)
    상업송장(wb)
    포장명세서(wb)
    wb.save(OUT)
    print(f'생성: {OUT}')


if __name__ == '__main__':
    main()
