#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""거래 데이터(JSON) → 상업송장 / 포장명세서.

forms/수출서류_상업송장_포장명세서.xlsx 템플릿의 [입력] 시트를 채우는 방식.
서식/수식/인쇄설정은 전부 템플릿이 갖고 있으므로 이 스크립트는 값만 넣는다.
템플릿 레이아웃을 고치려면 automation/build_template.py 를 고치고 다시 돌릴 것.

    python3 automation/generate.py automation/samples/sample.json -o out/
    python3 automation/generate.py <file.json> -o out/ --no-pdf

산출물
    {invoice_no}_수출서류.xlsx   [입력] + Commercial Invoice + Packing List (편집 가능)
    {invoice_no}_CI.pdf          Commercial Invoice 1장
    {invoice_no}_PL.pdf          Packing List 1장

PDF 는 LibreOffice(soffice)로 변환한다. 변환 시점에 수식이 계산되므로 별도 recalc 는 필요 없다.
soffice 가 없으면 xlsx 만 만들고 넘어간다.
"""
import argparse, json, os, re, shutil, subprocess, sys, tempfile

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
템플릿 = os.path.join(ROOT, 'forms', '수출서류_상업송장_포장명세서.xlsx')

품목_시작, 품목_끝 = 43, 57
품목_최대 = 품목_끝 - 품목_시작 + 1

# [입력] 시트 좌표 ← JSON 경로. build_template.py 의 입력시트() 와 1:1 대응.
거래_맵 = {
    'B5': 'invoice_no', 'B6': 'invoice_date', 'B7': 'po_no',
    'B8': '가격조건', 'B9': '인코텀즈버전', 'B10': '결제조건', 'B11': '통화',
    'B12': '원산지', 'B13': '선적항', 'B14': '도착항', 'B15': '최종목적지',
    'B16': '선박명', 'B17': '출항예정일',
}
당사자_맵 = {'상호': 0, '주소1': 1, '주소2': 2, '담당자': 3}
수출자_시작, 수입자_시작 = 20, 27          # 상호 행. +4 가 통관고유부호/해외거래처부호
품목_열 = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'I', 'J', 'K']   # H 는 수식이므로 건드리지 않음
품목_키 = ['_no', '품명', '규격', 'hs_code', '수량', '단위', '단가',
          '순중량_kg', '총중량_kg', '포장수량']

문서 = [('CI', 'COMMERCIAL INVOICE'), ('PL', 'PACKING LIST')]


def 채우기(d):
    if not os.path.exists(템플릿):
        sys.exit(f'템플릿이 없다: {템플릿}\n  python3 automation/build_template.py 로 먼저 생성할 것.')
    wb = openpyxl.load_workbook(템플릿)
    ws = wb['입력']
    거래 = d.get('거래', {})
    품목 = d.get('품목', [])

    if not 품목:
        sys.exit('품목이 비어 있다.')
    if len(품목) > 품목_최대:
        sys.exit(f'품목이 {len(품목)}개다. 템플릿은 {품목_최대}행까지다.\n'
                 f'  build_template.py 의 품목_끝 을 늘리고 다시 생성할 것.')

    for addr, key in 거래_맵.items():
        v = 거래.get(key)
        if v not in (None, ''):
            ws[addr] = v

    for 시작, 키, 부호키 in [(수출자_시작, '수출자', '통관고유부호'),
                            (수입자_시작, '수입자', '해외거래처부호')]:
        p = d.get(키, {}) or {}
        for 필드, off in 당사자_맵.items():
            v = p.get(필드)
            if v not in (None, ''):
                ws[f'B{시작 + off}'] = v
        v = p.get(부호키)
        if v not in (None, ''):
            ws[f'B{시작 + 4}'] = v

    ws['B34'] = d.get('착화통지처') or 'SAME AS CONSIGNEE'
    포장 = d.get('포장', {}) or {}
    ws['B37'] = 포장.get('포장단위') or 'CTNS'
    if 포장.get('총부피_CBM') is not None:
        ws['B38'] = 포장['총부피_CBM']

    # 품목 표: 템플릿의 예시 행을 지우고 실제 값으로 채운다 (H 열 수식은 유지)
    for i in range(품목_최대):
        r = 품목_시작 + i
        it = 품목[i] if i < len(품목) else None
        for col, key in zip(품목_열, 품목_키):
            if it is None:
                ws[f'{col}{r}'] = None
            elif key == '_no':
                ws[f'{col}{r}'] = i + 1
            else:
                ws[f'{col}{r}'] = it.get(key)
    return wb


def soffice():
    for name in ('soffice', 'libreoffice'):
        p = shutil.which(name)
        if p:
            return p
    return None


def 통째로PDF(xlsx, dst, timeout=300):
    """시트가 하나인 워크북을 그대로 PDF 로 바꾼다. 성공하면 페이지 수를 돌려준다."""
    exe = soffice()
    if not exe:
        print('  soffice 가 없어 PDF 를 건너뛴다. (apt-get install libreoffice-calc)')
        return None
    with tempfile.TemporaryDirectory() as tmp:
        r = subprocess.run([exe, '--headless', '-env:UserInstallation=file://' + tmp + '/profile',
                            '--convert-to', 'pdf', '--outdir', tmp, os.path.abspath(xlsx)],
                           capture_output=True, text=True, timeout=timeout)
        src = os.path.join(tmp, os.path.splitext(os.path.basename(xlsx))[0] + '.pdf')
        if not os.path.exists(src):
            print(f'  PDF 변환 실패: {(r.stderr or r.stdout).strip()[:200]}')
            return None
        shutil.copyfile(src, dst)
        if shutil.which('pdfinfo'):
            out = subprocess.run(['pdfinfo', dst], capture_output=True, text=True).stdout
            return int(out.split('Pages:')[1].split()[0])
        return 1


def pdf만들기(xlsx, outdir, invoice, timeout=300):
    """워크북 전체를 PDF 로 변환한 뒤, 시트별 제목으로 페이지를 갈라 문서별 PDF 를 만든다.

    시트를 숨겨도 LibreOffice PDF 변환에서는 제외되지 않으므로 페이지를 나누는 방식을 쓴다.
    """
    exe = soffice()
    if not exe:
        print('  soffice 가 없어 PDF 를 건너뛴다. (apt-get install libreoffice-calc)')
        return []
    if not shutil.which('pdftotext') or not shutil.which('pdfseparate'):
        print('  poppler-utils 가 없어 PDF 를 건너뛴다. (apt-get install poppler-utils)')
        return []

    with tempfile.TemporaryDirectory() as tmp:
        r = subprocess.run([exe, '--headless', '-env:UserInstallation=file://' + tmp + '/profile',
                            '--convert-to', 'pdf', '--outdir', tmp, os.path.abspath(xlsx)],
                           capture_output=True, text=True, timeout=timeout)
        전체 = os.path.join(tmp, os.path.splitext(os.path.basename(xlsx))[0] + '.pdf')
        if not os.path.exists(전체):
            print(f'  PDF 변환 실패: {(r.stderr or r.stdout).strip()[:200]}')
            return []

        n = int(subprocess.run(['pdfinfo', 전체], capture_output=True, text=True)
                .stdout.split('Pages:')[1].split()[0])

        # 각 페이지의 첫 줄에서 문서 제목을 찾아 페이지를 그룹으로 묶는다
        구간, 현재 = {}, None
        for p in range(1, n + 1):
            txt = subprocess.run(['pdftotext', '-f', str(p), '-l', str(p), 전체, '-'],
                                 capture_output=True, text=True).stdout
            머리 = [l.strip() for l in txt.splitlines() if l.strip()][:1]
            제목 = 머리[0] if 머리 else ''
            for 약칭, 표제 in 문서:
                if 제목 == 표제:
                    현재 = 약칭
                    구간.setdefault(약칭, [])
                    break
            if 현재:
                구간.setdefault(현재, []).append(p)

        만든것 = []
        for 약칭, 표제 in 문서:
            pages = 구간.get(약칭)
            if not pages:
                print(f'  {표제} 페이지를 찾지 못했다. 건너뛴다.')
                continue
            dst = os.path.join(outdir, f'{invoice}_{약칭}.pdf')
            if len(pages) == 1:
                subprocess.run(['pdfseparate', '-f', str(pages[0]), '-l', str(pages[0]),
                                전체, dst], check=True)
            else:
                조각 = os.path.join(tmp, f'{약칭}-%d.pdf')
                subprocess.run(['pdfseparate', '-f', str(pages[0]), '-l', str(pages[-1]),
                                전체, 조각], check=True)
                subprocess.run(['pdfunite'] + [os.path.join(tmp, f'{약칭}-{i}.pdf')
                                               for i in pages] + [dst], check=True)
            만든것.append((dst, len(pages)))
        return 만든것


def main():
    ap = argparse.ArgumentParser(description='거래 데이터 → 상업송장 / 포장명세서')
    ap.add_argument('데이터', help='거래 마스터 JSON (automation/schema.json 형식)')
    ap.add_argument('-o', '--out', default='.', help='출력 폴더')
    ap.add_argument('--no-pdf', action='store_true', help='xlsx 만 만들고 PDF 는 건너뛴다')
    ap.add_argument('--timeout', type=int, default=300, help='PDF 변환 제한시간(초)')
    a = ap.parse_args()

    with open(a.데이터, encoding='utf-8') as f:
        d = json.load(f)
    invoice = re.sub(r'[^\w.-]', '_', (d.get('거래', {}).get('invoice_no') or 'INVOICE'))
    os.makedirs(a.out, exist_ok=True)

    wb = 채우기(d)
    xlsx = os.path.join(a.out, f'{invoice}_수출서류.xlsx')
    wb.save(xlsx)
    print(f'생성: {xlsx}')
    print('       [입력] 시트를 고치면 두 서류가 함께 바뀐다.')

    if a.no_pdf:
        return 0
    for dst, pages in pdf만들기(xlsx, a.out, invoice, a.timeout):
        print(f'생성: {dst}  ({pages}장)')
    return 0


if __name__ == '__main__':
    sys.exit(main())
